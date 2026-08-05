"""
generate-products.py
Lee data/products.xlsx y genera data/products.json.
Ejecutar cada vez que el cliente actualice el Excel.

Uso:
  python generate-products.py
"""
import os, json
from datetime import datetime, timezone
from openpyxl import load_workbook

EXCEL_PATH  = 'data/products.xlsx'
JSON_PATH   = 'data/products.json'
IMAGES_JSON = 'data/product-images.json'
IMG_DIR     = 'img/productos'
FALLBACK    = 'img/placeholder.svg'

IMAGE_MAP = {
    '3870540': 'soft-gloves.png',
    '3870541': 'cling-wrap.png',
    '761765':  'liquid_detergent.png',
    '761710':  'liquid_detergent.png',
    '761758':  'cloth-softener.png',
    '761772':  'cloth-softener.png',
    '761703':  'dishwashing-liquid.png',
    '761789':  'dishwashing-liquid.png',
    '761796':  'floor-cleaner-pomegranate.png',
    '761734':  'floor-cleaner-pomegranate.png',
    '7907173': 'garbage-bag.png',
}

def resolve_image(sku_str):
    candidates = [
        os.path.join(IMG_DIR, f'{sku_str}.png'),
        os.path.join(IMG_DIR, f'{sku_str}.jpg'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path.replace('\\', '/'), 'SKU file'

    mapped = IMAGE_MAP.get(sku_str)
    if mapped:
        mapped_path = os.path.join(IMG_DIR, mapped)
        if os.path.exists(mapped_path):
            return mapped_path.replace('\\', '/'), f'map -> {mapped}'

    return FALLBACK, 'placeholder'

def load_image_overrides():
    """SKU -> lista de rutas (relativas al repo), la primera es la principal.
    Se genera desde tools/image-manager.py cuando un producto tiene varias fotos
    (ej: variantes de aroma/color, como Floor Cleaner)."""
    if not os.path.exists(IMAGES_JSON):
        return {}
    with open(IMAGES_JSON, encoding='utf-8') as f:
        return json.load(f)

def resolve_images(sku_str, overrides):
    paths = overrides.get(sku_str)
    if paths:
        valid = [p for p in paths if os.path.exists(p)]
        if valid:
            return valid, 'product-images.json'
    single, how = resolve_image(sku_str)
    return [single], how

# ── Leer Excel ────────────────────────────────────────────────────────────
if not os.path.exists(EXCEL_PATH):
    print(f'ERROR: {EXCEL_PATH} not found. Run create_products_excel.py first.')
    raise SystemExit(1)

wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Products']

overrides = load_image_overrides()
products  = []
skipped   = 0

for row in ws.iter_rows(min_row=5, values_only=True):
    sku, name, category, unit_price, box_price, box_contents, available = row

    if not sku:                          # fila vacía
        continue
    if str(available).strip().upper() != 'Y':
        skipped += 1
        continue

    sku_str        = str(int(float(str(sku).replace(',', ''))))
    images, _how   = resolve_images(sku_str, overrides)

    if unit_price is None or box_price is None or float(box_price) == 0:
        print(f'  ERROR: SKU {sku_str} skipped — missing or zero unit/box price')
        skipped += 1
        continue
    unit  = round(float(unit_price),  2)
    box   = round(float(box_price),   2)

    product = {
        'id':            sku_str,
        'name':          str(name).strip(),
        'category':      str(category).strip().lower(),
        'categoryLabel': str(category).strip(),
        'unitPrice':     unit,
        'unitCents':     int(round(unit  * 100)),
        'boxPrice':      box,
        'boxPriceCents': int(round(box   * 100)),
        'boxContents':   str(box_contents).strip(),
        'image':         images[0],
        'imageSource':   _how,
        'available':     True,
    }
    if len(images) > 1:
        product['images'] = images
    products.append(product)

# ── Categorías únicas ordenadas ───────────────────────────────────────────
categories = sorted({p['categoryLabel'] for p in products})

output = {
    'generated':  datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
    'total':      len(products),
    'categories': categories,
    'products':   products,
}

# ── Escribir JSON ─────────────────────────────────────────────────────────
os.makedirs('data', exist_ok=True)
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

# ── Reporte ───────────────────────────────────────────────────────────────
print()
print(f'OK  {len(products)} products exported to {JSON_PATH}')
if skipped:
    print(f'   ({skipped} products skipped — Available = N)')
print()
print(f'  {"SKU":<10} {"Name":<40} {"Unit":>7}  {"Box":>8}  Image')
print(f'  {"-"*10} {"-"*40} {"-"*7}  {"-"*8}  {"-"*30}')
for p in products:
    img_label = p['imageSource'] if p['image'] != FALLBACK else 'PLACEHOLDER'
    print(f'  {p["id"]:<10} {p["name"][:40]:<40} ${p["unitPrice"]:>6.2f}  ${p["boxPrice"]:>7.2f}  {img_label}')

print()
print('Next steps:')
print('  Upload to GitHub:  data/products.json')
print('  (Optional)         data/products.xlsx')
print('  Add product images: img/productos/{SKU}.png  (e.g. 761765.png)')
