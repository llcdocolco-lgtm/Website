"""
Docolco Product Manager — tools/image-manager.py
Gestión de imágenes y catálogo de productos.
Doble clic en "Iniciar Image Manager.bat" para abrir.
"""

import os
import sys
import json
import subprocess
import threading
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

try:
    from PIL import Image, ImageTk
    PILLOW_OK = True
except ImportError:
    PILLOW_OK = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

REPO_ROOT       = Path(__file__).parent.parent.resolve()
IMG_DIR         = REPO_ROOT / "img" / "productos"
GENERATE_SCRIPT = REPO_ROOT / "generate-products.py"
EXCEL_PATH      = REPO_ROOT / "data" / "products.xlsx"
IMAGES_JSON     = REPO_ROOT / "data" / "product-images.json"
SITE_URL        = "https://docolco.netlify.app"

BLUE     = "#1A3FA8"
BLUE_DRK = "#122E80"
WHITE    = "#FFFFFF"
LIGHT    = "#F8F8F6"
CARD_BG  = "#FFFFFF"
BORDER_C = "#E4E4E0"
MUTED    = "#888888"
GREEN    = "#1B5E20"
RED      = "#CC0000"

CATEGORIES    = ["Cleaning", "Packaging", "Protection", "Waste"]
ACCEPTED_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".bmp", ".tiff", ".tif"}

# Mapa de imágenes con nombre propio para SKUs que no tienen archivo {sku}.jpg/png
# (debe coincidir con generate-products.py para que la miniatura sea la real).
IMAGE_MAP = {
    "3870540": "soft-gloves.png",
    "3870541": "cling-wrap.png",
    "761765":  "liquid_detergent.png",
    "761710":  "liquid_detergent.png",
    "761758":  "cloth-softener.png",
    "761772":  "cloth-softener.png",
    "761703":  "dishwashing-liquid.png",
    "761789":  "dishwashing-liquid.png",
    "761796":  "floor-cleaner-pomegranate.png",
    "761734":  "floor-cleaner-pomegranate.png",
    "7907173": "garbage-bag.png",
}

if OPENPYXL_OK:
    _XL_GRAY      = PatternFill("solid", fgColor="F0F0F0")
    _XL_YELLOW    = PatternFill("solid", fgColor="FFFDE7")
    _XL_THIN      = Border(bottom=Side(style="thin", color="DDDDDD"))
    _XL_CENTER    = Alignment(horizontal="center", vertical="center")
    _XL_LEFT      = Alignment(horizontal="left",   vertical="center")
    _XL_NORM      = Font(name="Calibri", size=10)
    _XL_SKU_F     = Font(name="Calibri", color="555555", size=10)
    _XL_PRICE_FMT = "$#,##0.00"

CARD_W, CARD_H   = 168, 208
THUMB_W, THUMB_H = 148, 110
PREVIEW_W, PREVIEW_H = 260, 190


def resolve_product_image(sku):
    """Devuelve la ruta de la imagen real de un SKU, o None si no hay ninguna."""
    for ext in (".jpg", ".jpeg", ".png", ".webp"):
        p = IMG_DIR / f"{sku}{ext}"
        if p.exists():
            return p
    mapped = IMAGE_MAP.get(str(sku))
    if mapped:
        p = IMG_DIR / mapped
        if p.exists():
            return p
    return None


def load_image_overrides():
    """SKU -> lista de rutas relativas al repo (la primera es la principal)."""
    if not IMAGES_JSON.exists():
        return {}
    try:
        return json.loads(IMAGES_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_image_overrides(overrides):
    IMAGES_JSON.parent.mkdir(parents=True, exist_ok=True)
    IMAGES_JSON.write_text(
        json.dumps(overrides, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def resolve_images_for_sku(sku, overrides=None):
    """Lista de rutas (Path) de todas las fotos de un SKU. La primera es la principal.
    Usa data/product-images.json si el producto tiene varias fotos asignadas;
    si no, cae al archivo {sku}.jpg/png de siempre."""
    if overrides is None:
        overrides = load_image_overrides()
    rel_paths = overrides.get(str(sku))
    if rel_paths:
        valid = [REPO_ROOT / rp for rp in rel_paths if (REPO_ROOT / rp).exists()]
        if valid:
            return valid
    single = resolve_product_image(sku)
    return [single] if single else []


class ImageManagerApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Docolco Product Manager")
        self.geometry("900x720")
        self.resizable(True, True)
        self.minsize(760, 560)
        self.configure(bg=WHITE)

        self._publishing   = False
        self._products      = self._read_excel_products()
        self._thumb_cache   = {}   # sku -> PhotoImage (evita garbage collection)
        self._search_var    = tk.StringVar()

        self._build_menu()
        self._build_ui()
        self.after(400, self._startup_check)

    # ── Data layer ────────────────────────────────────────────────────────────

    def _read_excel_products(self):
        if not OPENPYXL_OK or not EXCEL_PATH.exists():
            return []
        try:
            wb = load_workbook(EXCEL_PATH, data_only=True)
            ws = wb["Products"]
            products = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                sku = row[0]
                if not sku:
                    continue
                _, name, category, unit_price, box_price, box_contents, available = row
                sku_str = str(int(float(str(sku).replace(",", ""))))
                products.append({
                    "sku":          sku_str,
                    "name":         str(name or "").strip(),
                    "category":     str(category or "").strip(),
                    "unit_price":   round(float(unit_price or 0), 2),
                    "box_price":    round(float(box_price or 0), 2),
                    "box_contents": str(box_contents or "").strip(),
                    "available":    str(available or "Y").strip().upper(),
                })
            return products
        except Exception:
            return []

    def _write_excel_products(self, products):
        if not OPENPYXL_OK:
            raise RuntimeError("openpyxl no está instalado.\nCorre tools/setup.bat primero.")
        if not EXCEL_PATH.exists():
            raise RuntimeError(f"No se encontró data/products.xlsx.\nEsperado en:\n{EXCEL_PATH}")
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Products"]
        for row_num in range(5, ws.max_row + 1):
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).value = None
        for i, p in enumerate(products):
            rn = i + 5
            vals = [
                int(p["sku"]) if str(p["sku"]).isdigit() else p["sku"],
                p["name"], p["category"],
                p["unit_price"], p["box_price"],
                p["box_contents"], p["available"],
            ]
            for ci, val in enumerate(vals, start=1):
                cell = ws.cell(row=rn, column=ci, value=val)
                cell.border    = _XL_THIN
                cell.alignment = _XL_CENTER if ci in (1, 3, 4, 5, 7) else _XL_LEFT
                if ci == 1:
                    cell.fill = _XL_GRAY
                    cell.font = _XL_SKU_F
                elif ci in (4, 5):
                    cell.fill          = _XL_YELLOW
                    cell.font          = _XL_NORM
                    cell.number_format = _XL_PRICE_FMT
                else:
                    cell.fill = _XL_YELLOW
                    cell.font = _XL_NORM
            ws.row_dimensions[rn].height = 30
        wb.save(EXCEL_PATH)

    def _current_branch(self):
        try:
            r = subprocess.run(
                ["git", "rev-parse", "--abbrev-ref", "HEAD"],
                capture_output=True, text=True, cwd=str(REPO_ROOT), check=True,
            )
            return r.stdout.strip() or "main"
        except Exception:
            return "main"

    # ── Menu ──────────────────────────────────────────────────────────────────

    def _build_menu(self):
        bar = tk.Menu(self)
        self.config(menu=bar)

        file_m = tk.Menu(bar, tearoff=0)
        file_m.add_command(label="Abrir carpeta del repo",
                           command=lambda: os.startfile(str(REPO_ROOT)))
        file_m.add_command(label="Abrir Excel de productos",
                           command=self._open_excel)
        bar.add_cascade(label="Archivo", menu=file_m)

        tools_m = tk.Menu(bar, tearoff=0)
        tools_m.add_command(label="Ver imágenes actuales", command=self._show_images)
        tools_m.add_command(label="Verificar estado Git",  command=self._verify_status)
        bar.add_cascade(label="Herramientas", menu=tools_m)

        help_m = tk.Menu(bar, tearoff=0)
        help_m.add_command(label="Contactar soporte", command=self._contact_support)
        bar.add_cascade(label="Ayuda", menu=help_m)

    # ── Main UI ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        hdr = tk.Frame(self, bg=BLUE, height=70)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="DOCOLCO", bg=BLUE, fg=WHITE,
                 font=("Helvetica", 18, "bold")).pack(anchor="w", padx=20, pady=(12, 0))
        tk.Label(hdr, text="Product Manager", bg=BLUE, fg="#99B8FF",
                 font=("Helvetica", 10)).pack(anchor="w", padx=20)

        toolbar = tk.Frame(self, bg=WHITE)
        toolbar.pack(fill="x", padx=16, pady=(12, 4))

        tk.Button(
            toolbar, text="+ Agregar producto",
            bg=BLUE, fg=WHITE, relief="flat", bd=0, cursor="hand2",
            font=("Helvetica", 10, "bold"), padx=12, pady=6,
            command=lambda: self._open_product_dialog(None),
        ).pack(side="left")

        search_entry = tk.Entry(toolbar, textvariable=self._search_var,
                                 font=("Helvetica", 10), relief="solid", bd=1)
        search_entry.pack(side="right", ipady=4, padx=(6, 0))
        search_entry.insert(0, "")
        tk.Label(toolbar, text="Buscar:", bg=WHITE, fg=MUTED,
                 font=("Helvetica", 9)).pack(side="right")
        self._search_var.trace_add("write", lambda *_: self._render_grid())

        # Placeholder para el hint "producto agregado" tipo texto guía
        tk.Label(
            self, text="Clic en un producto para ver su ficha completa, editar el texto o cambiar la imagen.",
            bg=WHITE, fg=MUTED, font=("Helvetica", 9),
        ).pack(anchor="w", padx=18, pady=(0, 4))

        # ── Área con scroll ─────────────────────────────────────────────────
        outer = tk.Frame(self, bg=WHITE)
        outer.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        self._canvas = tk.Canvas(outer, bg=LIGHT, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        self._canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self._grid_frame = tk.Frame(self._canvas, bg=LIGHT)
        self._grid_window = self._canvas.create_window((0, 0), window=self._grid_frame, anchor="nw")

        self._grid_frame.bind("<Configure>", lambda _e: self._canvas.configure(
            scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>", self._on_canvas_resize)
        self._canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        self._render_grid()

    def _on_canvas_resize(self, event):
        self._canvas.itemconfig(self._grid_window, width=event.width)
        self._render_grid()

    def _on_mousewheel(self, event):
        self._canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ── Grid de productos ────────────────────────────────────────────────────

    def _render_grid(self):
        for w in self._grid_frame.winfo_children():
            w.destroy()

        query = self._search_var.get().strip().lower()
        items = [
            p for p in self._products
            if not query or query in p["name"].lower() or query in p["sku"]
        ]

        if not items:
            tk.Label(
                self._grid_frame,
                text="Sin resultados." if query else "No hay productos todavía. Usa '+ Agregar producto'.",
                bg=LIGHT, fg=MUTED, font=("Helvetica", 10), pady=30,
            ).grid(row=0, column=0, sticky="w", padx=8)
            return

        canvas_w = max(self._canvas.winfo_width(), CARD_W)
        cols = max(1, canvas_w // (CARD_W + 16))

        for col in range(cols):
            self._grid_frame.grid_columnconfigure(col, weight=1)

        for i, p in enumerate(items):
            row, col = divmod(i, cols)
            self._build_card(self._grid_frame, p, row, col)

    def _build_card(self, parent, product, row, col):
        card = tk.Frame(parent, bg=CARD_BG, width=CARD_W, height=CARD_H,
                         highlightbackground=BORDER_C, highlightthickness=1, cursor="hand2")
        card.grid(row=row, column=col, padx=8, pady=8, sticky="n")
        card.grid_propagate(False)
        card.pack_propagate(False)

        img_holder = tk.Frame(card, bg=CARD_BG, width=THUMB_W, height=THUMB_H)
        img_holder.pack(pady=(10, 6))
        img_holder.pack_propagate(False)

        photo = self._get_thumbnail(product["sku"])
        if photo:
            img_lbl = tk.Label(img_holder, image=photo, bg=CARD_BG)
        else:
            img_lbl = tk.Label(img_holder, text="Sin imagen", bg=LIGHT, fg=MUTED,
                                font=("Helvetica", 9))
            img_lbl.pack(fill="both", expand=True)
        if photo:
            img_lbl.pack()

        name_lbl = tk.Label(card, text=product["name"] or "(sin nombre)", bg=CARD_BG,
                             font=("Helvetica", 9, "bold"), wraplength=CARD_W - 16,
                             justify="center")
        name_lbl.pack(padx=6)

        price_lbl = tk.Label(card, text=f"${product['unit_price']:.2f}", bg=CARD_BG,
                              fg=BLUE, font=("Helvetica", 9))
        price_lbl.pack(pady=(2, 0))

        badge_text = "Activo" if product["available"] == "Y" else "Inactivo"
        badge_fg   = GREEN if product["available"] == "Y" else MUTED
        tk.Label(card, text=badge_text, bg=CARD_BG, fg=badge_fg,
                 font=("Helvetica", 8)).pack(pady=(2, 6))

        for w in (card, img_holder, img_lbl, name_lbl, price_lbl):
            w.bind("<Button-1>", lambda _e, sku=product["sku"]: self._open_product_dialog(sku))

    def _get_thumbnail(self, sku):
        if sku in self._thumb_cache:
            return self._thumb_cache[sku]
        if not PILLOW_OK:
            return None
        images = resolve_images_for_sku(sku)
        path = images[0] if images else None
        if not path:
            return None
        try:
            img = Image.open(path)
            img.thumbnail((THUMB_W, THUMB_H), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self._thumb_cache[sku] = photo
            return photo
        except Exception:
            return None

    # ── Diálogo de ficha de producto (crear / editar) ────────────────────────

    def _open_product_dialog(self, sku):
        is_new  = sku is None
        product = None
        if not is_new:
            product = next((p for p in self._products if p["sku"] == sku), None)
            if product is None:
                return

        dlg = tk.Toplevel(self)
        dlg.title("Nuevo producto" if is_new else f"Producto — {product['name']}")
        dlg.geometry("620x560")
        dlg.resizable(False, False)
        dlg.configure(bg=WHITE)
        dlg.grab_set()

        state = {"images": [], "publishing": False}
        if not is_new:
            state["images"] = [{"path": p, "is_new": False}
                                for p in resolve_images_for_sku(sku)]

        body = tk.Frame(dlg, bg=WHITE)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # ── Columna izquierda: fotos (varias por producto) ───────────────────
        left = tk.Frame(body, bg=WHITE, width=PREVIEW_W + 20)
        left.pack(side="left", fill="y", padx=(0, 18))
        left.pack_propagate(False)

        tk.Label(left, text="Fotos del producto", bg=WHITE,
                 font=("Helvetica", 10, "bold"), anchor="w").pack(fill="x", pady=(0, 2))
        tk.Label(left, text="La primera es la principal en el sitio. "
                            "Clic en una miniatura para hacerla principal "
                            "(igual que las variantes de Floor Cleaner).",
                 bg=WHITE, fg=MUTED, font=("Helvetica", 8), anchor="w",
                 wraplength=PREVIEW_W, justify="left").pack(fill="x", pady=(0, 6))

        zone = tk.Frame(left, bg=LIGHT, width=PREVIEW_W, height=PREVIEW_H,
                         highlightbackground=BLUE, highlightthickness=2, cursor="hand2")
        zone.pack()
        zone.pack_propagate(False)

        preview_lbl = tk.Label(zone, bg=LIGHT, cursor="hand2")
        preview_lbl.pack(fill="both", expand=True)

        thumbs_frame = tk.Frame(left, bg=WHITE)
        thumbs_frame.pack(fill="x", pady=(8, 0))
        thumb_photos = []   # referencias vivas para que Tk no las recolecte

        def _set_preview(path):
            if not PILLOW_OK or not path:
                preview_lbl.config(image="", text="Sin fotos todavía\nclic para agregar una",
                                    fg=MUTED, font=("Helvetica", 10))
                preview_lbl.image = None
                return
            try:
                img = Image.open(path)
                img.thumbnail((PREVIEW_W, PREVIEW_H), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                preview_lbl.image = photo   # referencia viva
                preview_lbl.config(image=photo, text="")
            except Exception:
                preview_lbl.config(image="", text="No se pudo\nleer la imagen",
                                    fg=RED, font=("Helvetica", 10))

        def _make_thumb_photo(path):
            if not PILLOW_OK:
                return None
            try:
                img = Image.open(path)
                img.thumbnail((56, 44), Image.LANCZOS)
                return ImageTk.PhotoImage(img)
            except Exception:
                return None

        def _make_primary(i):
            if i == 0:
                return
            state["images"].insert(0, state["images"].pop(i))
            _refresh_gallery()

        def _remove_image(i):
            state["images"].pop(i)
            _refresh_gallery()

        def _refresh_gallery():
            thumb_photos.clear()
            for w in thumbs_frame.winfo_children():
                w.destroy()
            _set_preview(state["images"][0]["path"] if state["images"] else None)
            for idx, item in enumerate(state["images"]):
                cell = tk.Frame(thumbs_frame, bg=WHITE)
                cell.pack(side="left", padx=(0, 6))
                photo = _make_thumb_photo(item["path"])
                if photo:
                    thumb_photos.append(photo)
                    thumb_lbl = tk.Label(
                        cell, image=photo, bg=WHITE, cursor="hand2",
                        highlightthickness=2,
                        highlightbackground=BLUE if idx == 0 else BORDER_C,
                    )
                else:
                    thumb_lbl = tk.Label(cell, text="?", bg=LIGHT, width=8, height=3,
                                         cursor="hand2")
                thumb_lbl.pack()
                thumb_lbl.bind("<Button-1>", lambda _e, i=idx: _make_primary(i))
                tk.Label(
                    cell, text=("Principal" if idx == 0 else "Hacer principal"),
                    bg=WHITE, fg=(GREEN if idx == 0 else BLUE), cursor="hand2",
                    font=("Helvetica", 7),
                ).pack()
                rm_lbl = tk.Label(cell, text="Quitar", bg=WHITE, fg=RED, cursor="hand2",
                                   font=("Helvetica", 7, "underline"))
                rm_lbl.pack()
                rm_lbl.bind("<Button-1>", lambda _e, i=idx: _remove_image(i))

        def _add_images():
            paths = filedialog.askopenfilenames(
                title="Seleccionar imagen(es) del producto",
                filetypes=[
                    ("Imágenes", "*.jpg *.jpeg *.png *.webp *.heic *.bmp *.tiff *.tif"),
                    ("Todos los archivos", "*.*"),
                ],
            )
            if not paths:
                return
            for raw in paths:
                p = Path(raw)
                if p.suffix.lower() not in ACCEPTED_EXTS:
                    messagebox.showerror(
                        "Formato no compatible",
                        f"'{p.name}' no es compatible.\nUsa: JPG, PNG, WEBP, HEIC, BMP o TIFF.",
                    )
                    continue
                state["images"].append({"path": p, "is_new": True})
            _refresh_gallery()

        zone.bind("<Button-1>", lambda _e: _add_images())
        preview_lbl.bind("<Button-1>", lambda _e: _add_images())
        tk.Button(left, text="+ Agregar imagen", command=_add_images,
                  bg=LIGHT, fg=BLUE, relief="flat", bd=1, cursor="hand2",
                  font=("Helvetica", 9), pady=4).pack(fill="x", pady=(8, 0))

        _refresh_gallery()

        # ── Columna derecha: campos de texto ─────────────────────────────────
        right = tk.Frame(body, bg=WHITE)
        right.pack(side="left", fill="both", expand=True)

        fields = {}

        def _field(label, key, default="", combo_opts=None):
            tk.Label(right, text=label, bg=WHITE,
                     font=("Helvetica", 10, "bold"), anchor="w").pack(fill="x")
            var = tk.StringVar(value=default)
            if combo_opts:
                w = ttk.Combobox(right, textvariable=var, values=combo_opts,
                                  state="readonly", font=("Helvetica", 10))
            else:
                w = tk.Entry(right, textvariable=var, font=("Helvetica", 10))
            w.pack(fill="x", pady=(2, 10))
            fields[key] = var
            return w

        if is_new:
            sku_entry = _field("SKU  *", "sku")
            sku_entry.focus_set()
        else:
            tk.Label(right, text="SKU", bg=WHITE, font=("Helvetica", 10, "bold"),
                     anchor="w").pack(fill="x")
            tk.Label(right, text=product["sku"], bg=WHITE, fg=MUTED,
                     font=("Helvetica", 10), anchor="w").pack(fill="x", pady=(2, 10))
            fields["sku"] = tk.StringVar(value=product["sku"])

        _field("Nombre del producto  *", "name", product["name"] if product else "")
        _field("Categoría  *", "category",
               product["category"] if product else "Cleaning", CATEGORIES)
        _field("Precio unitario ($)  *", "unit_price",
               f"{product['unit_price']:.2f}" if product else "0.00")
        _field("Precio por caja ($)  *", "box_price",
               f"{product['box_price']:.2f}" if product else "0.00")
        _field("Contenido por caja  *", "box_contents",
               product["box_contents"] if product else "12 Units/Box")
        _field("Activo (Y/N)", "available",
               product["available"] if product else "Y", ["Y", "N"])

        err_lbl = tk.Label(right, text="", bg=WHITE, fg=RED, font=("Helvetica", 9),
                            wraplength=320, justify="left")
        err_lbl.pack(fill="x")

        log_wrap = tk.Frame(right, bg=WHITE)
        log_wrap.pack(fill="x", pady=(6, 0))
        log_txt = tk.Text(log_wrap, height=4, font=("Courier", 9), bg=LIGHT,
                           relief="flat", bd=1, wrap="word", state="disabled")
        log_txt.pack(fill="x")

        def _log(msg):
            def _do():
                log_txt.config(state="normal")
                log_txt.insert("end", msg + "\n")
                log_txt.see("end")
                log_txt.config(state="disabled")
            dlg.after(0, _do)

        # ── Botonera ──────────────────────────────────────────────────────────
        btn_row = tk.Frame(dlg, bg=WHITE)
        btn_row.pack(fill="x", side="bottom", padx=20, pady=(0, 16))

        save_btn = tk.Button(
            btn_row, text="Guardar y Publicar",
            bg=GREEN, fg=WHITE, relief="flat", bd=0, cursor="hand2",
            font=("Helvetica", 10, "bold"), padx=14, pady=7,
        )
        save_btn.pack(side="right")

        tk.Button(btn_row, text="Cancelar", command=dlg.destroy,
                  bg=LIGHT, fg=MUTED, relief="flat", bd=1, cursor="hand2",
                  font=("Helvetica", 10), padx=12, pady=7).pack(side="right", padx=(0, 8))

        if not is_new:
            tk.Button(
                btn_row, text="Eliminar producto",
                command=lambda: self._confirm_delete(dlg, product, _log, save_btn),
                bg=WHITE, fg=RED, relief="flat", bd=1, cursor="hand2",
                font=("Helvetica", 10), padx=12, pady=7,
            ).pack(side="left")

        def _collect_and_validate():
            data = {k: v.get().strip() for k, v in fields.items()}
            if not data["sku"]:
                err_lbl.config(text="El SKU es obligatorio.")
                return None
            if not data["name"]:
                err_lbl.config(text="El nombre es obligatorio.")
                return None
            if is_new and any(p["sku"] == data["sku"] for p in self._products):
                err_lbl.config(text=f"El SKU '{data['sku']}' ya existe.")
                return None
            try:
                up = round(float(data["unit_price"].replace("$", "").replace(",", "")), 2)
                bp = round(float(data["box_price"].replace("$", "").replace(",", "")), 2)
            except ValueError:
                err_lbl.config(text="Los precios deben ser números (ej: 1.99).")
                return None
            err_lbl.config(text="")
            return {
                "sku":          data["sku"],
                "name":         data["name"],
                "category":     data["category"] or "Cleaning",
                "unit_price":   up,
                "box_price":    bp,
                "box_contents": data["box_contents"],
                "available":    data["available"] or "Y",
            }

        def _on_save():
            if state["publishing"]:
                return
            new_data = _collect_and_validate()
            if new_data is None:
                return
            state["publishing"] = True
            save_btn.config(state="disabled", text="Publicando...")
            threading.Thread(
                target=self._publish_product_worker,
                args=(new_data, list(state["images"]), is_new, sku, dlg, state, save_btn, _log),
                daemon=True,
            ).start()

        save_btn.config(command=_on_save)
        dlg.bind("<Return>", lambda _e: _on_save())

    def _confirm_delete(self, dlg, product, _log, save_btn):
        if not messagebox.askyesno(
            "Confirmar eliminación",
            f"¿Eliminar '{product['name']}' (SKU {product['sku']})?\n\n"
            "El producto dejará de aparecer en el sitio.",
            parent=dlg,
        ):
            return
        save_btn.config(state="disabled")
        threading.Thread(
            target=self._delete_product_worker,
            args=(product["sku"], dlg, _log),
            daemon=True,
        ).start()

    # ── Workers (hilo de fondo) ───────────────────────────────────────────────

    def _git_sync_and_commit(self, paths, commit_msg, _log):
        branch = self._current_branch()
        _log("Sincronizando con GitHub...")
        pull = subprocess.run(
            ["git", "pull", "--rebase", "origin", branch],
            capture_output=True, text=True, cwd=str(REPO_ROOT),
        )
        if pull.returncode != 0:
            subprocess.run(["git", "rebase", "--abort"],
                           capture_output=True, cwd=str(REPO_ROOT))
            _log(f"  (aviso sync: {pull.stderr.strip()[:80]})")

        subprocess.run(["git", "add", *paths], capture_output=True,
                       cwd=str(REPO_ROOT), check=True)
        commit = subprocess.run(
            ["git", "commit", "-m", commit_msg],
            capture_output=True, text=True, cwd=str(REPO_ROOT),
        )
        commit_out = (commit.stdout + commit.stderr).lower()
        if commit.returncode != 0 and "nothing to commit" not in commit_out:
            raise RuntimeError(f"git commit falló:\n{commit.stderr}")
        _log("✓ Commit creado")

        push = subprocess.run(
            ["git", "push", "origin", branch],
            capture_output=True, text=True, cwd=str(REPO_ROOT),
        )
        if push.returncode != 0:
            raise RuntimeError(
                "Sin conexión a internet. Los cambios se guardaron localmente.\n"
                "Intenta publicar de nuevo cuando tengas conexión.\n\n"
                f"Detalle:\n{push.stderr.strip()[:200]}"
            )
        _log(f"✓ ¡Publicado! → {SITE_URL}")

    def _publish_product_worker(self, data, images, is_new, old_sku, dlg, state, save_btn, _log):
        try:
            changed_paths = ["data/products.json", "data/product-images.json"]

            final_paths = []
            if images:
                IMG_DIR.mkdir(parents=True, exist_ok=True)
                count = len(images)
                for idx, item in enumerate(images):
                    if not item["is_new"]:
                        final_paths.append(item["path"])
                        continue
                    if not PILLOW_OK:
                        raise RuntimeError("Pillow no está instalado.\nCorre tools/setup.bat primero.")
                    name = f"{data['sku']}.jpg" if count == 1 else f"{data['sku']}-{idx + 1}.jpg"
                    dest = IMG_DIR / name
                    _log(f"Convirtiendo {item['path'].name} a JPG...")
                    img = Image.open(item["path"])
                    if img.mode in ("RGBA", "LA", "P"):
                        background = Image.new("RGB", img.size, (234, 240, 255))
                        if img.mode == "P":
                            img = img.convert("RGBA")
                        if img.mode in ("RGBA", "LA"):
                            background.paste(img, mask=img.getchannel("A"))
                        else:
                            background.paste(img)
                        img = background
                    elif img.mode != "RGB":
                        img = img.convert("RGB")
                    img.save(dest, "JPEG", quality=90)
                    changed_paths.append(dest.relative_to(REPO_ROOT).as_posix())
                    final_paths.append(dest)
                    _log(f"✓ Guardada como {name}")
                self._thumb_cache.pop(data["sku"], None)

            overrides = load_image_overrides()
            if final_paths:
                overrides[data["sku"]] = [p.relative_to(REPO_ROOT).as_posix() for p in final_paths]
            else:
                overrides.pop(data["sku"], None)
            save_image_overrides(overrides)

            if is_new:
                self._products.append(data)
            else:
                for i, p in enumerate(self._products):
                    if p["sku"] == old_sku:
                        self._products[i] = data
                        break

            _log("Guardando Excel...")
            self._write_excel_products(self._products)
            _log("✓ products.xlsx actualizado")

            _log("Regenerando catálogo JSON...")
            r = subprocess.run(
                [sys.executable, str(GENERATE_SCRIPT)],
                capture_output=True, text=True, cwd=str(REPO_ROOT),
            )
            if r.returncode != 0:
                raise RuntimeError(f"generate-products.py falló:\n{r.stderr or r.stdout}")
            _log("✓ products.json actualizado")

            commit_msg = (f"feat: agregar producto {data['sku']}" if is_new
                          else f"feat: actualizar producto {data['sku']}")
            self._git_sync_and_commit(changed_paths, commit_msg, _log)

            self.after(0, self._on_product_success, dlg)

        except FileNotFoundError as exc:
            msg = self._git_not_found_msg(exc)
            self.after(0, self._on_product_error, msg, state, save_btn)
        except Exception as exc:
            self.after(0, self._on_product_error, str(exc), state, save_btn)

    def _delete_product_worker(self, sku, dlg, _log):
        try:
            self._products = [p for p in self._products if p["sku"] != sku]

            overrides = load_image_overrides()
            if sku in overrides:
                del overrides[sku]
                save_image_overrides(overrides)
            self._thumb_cache.pop(sku, None)

            _log("Guardando Excel...")
            self._write_excel_products(self._products)
            _log("✓ products.xlsx actualizado")

            _log("Regenerando catálogo JSON...")
            r = subprocess.run(
                [sys.executable, str(GENERATE_SCRIPT)],
                capture_output=True, text=True, cwd=str(REPO_ROOT),
            )
            if r.returncode != 0:
                raise RuntimeError(f"generate-products.py falló:\n{r.stderr or r.stdout}")
            _log("✓ products.json actualizado")

            self._git_sync_and_commit(
                ["data/products.json", "data/product-images.json"],
                f"fix: eliminar producto {sku}", _log,
            )
            self.after(0, self._on_product_success, dlg)

        except FileNotFoundError as exc:
            msg = self._git_not_found_msg(exc)
            self.after(0, self._on_product_error, msg, {"publishing": False}, None)
        except Exception as exc:
            self.after(0, self._on_product_error, str(exc), {"publishing": False}, None)

    def _git_not_found_msg(self, exc):
        if "git" in str(exc).lower():
            return (
                "Git no está instalado en este equipo.\n\n"
                "Descarga e instala Git para Windows (git-scm.com)\n"
                "y reinicia la app.\n\nContacta a Samuel:\n+57 304 353 8450"
            )
        return f"Archivo no encontrado:\n{exc}"

    def _on_product_success(self, dlg):
        self._render_grid()
        try:
            dlg.destroy()
        except tk.TclError:
            pass
        messagebox.showinfo(
            "¡Publicado!",
            f"Los cambios están en línea.\n\nEl sitio se actualiza en 1-2 minutos:\n{SITE_URL}",
        )

    def _on_product_error(self, msg, state, save_btn):
        state["publishing"] = False
        if save_btn is not None:
            try:
                save_btn.config(state="normal", text="Guardar y Publicar")
            except tk.TclError:
                pass
        messagebox.showerror("Error al publicar", msg)

    # ── Menu handlers ─────────────────────────────────────────────────────────

    def _open_excel(self):
        if not EXCEL_PATH.exists():
            messagebox.showinfo("Excel",
                                f"No se encontró data/products.xlsx.\nEsperado en:\n{EXCEL_PATH}")
            return
        try:
            os.startfile(str(EXCEL_PATH))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el Excel:\n{e}")

    def _show_images(self):
        if not IMG_DIR.exists():
            messagebox.showinfo("Imágenes", "La carpeta img/productos/ no existe aún.")
            return
        imgs = sorted(IMG_DIR.iterdir())
        if not imgs:
            messagebox.showinfo("Imágenes", "No hay imágenes en img/productos/")
            return
        used_names = set()
        for p in self._products:
            for path in resolve_images_for_sku(p["sku"]):
                used_names.add(path.name)
        lines = [
            ("✓" if i.name in used_names else "○") + f"  {i.name}"
            + ("  ← activo" if i.name in used_names else "")
            for i in imgs
        ]
        win = tk.Toplevel(self)
        win.title("Imágenes actuales")
        win.geometry("440x300")
        win.configure(bg=WHITE)
        tk.Label(win,
                 text="✓ = asignada a un producto activo   ○ = no usada por ningún producto",
                 bg=WHITE, fg=MUTED, font=("Helvetica", 9)).pack(anchor="w", padx=12, pady=(10, 4))
        t = tk.Text(win, font=("Courier", 9), bg=LIGHT, wrap="none")
        t.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        t.insert("1.0", "\n".join(lines))
        t.config(state="disabled")

    def _verify_status(self):
        try:
            r = subprocess.run(["git", "status", "--short"],
                               capture_output=True, text=True,
                               cwd=str(REPO_ROOT), check=True)
            out = r.stdout.strip() or "✓ Todo sincronizado con GitHub."
        except FileNotFoundError:
            out = "✗ Git no está instalado."
        except Exception as e:
            out = str(e)
        win = tk.Toplevel(self)
        win.title("Estado del repositorio")
        win.geometry("400x220")
        win.configure(bg=WHITE)
        t = tk.Text(win, font=("Courier", 9), bg=LIGHT, height=8)
        t.pack(fill="both", expand=True, padx=12, pady=12)
        t.insert("1.0", out)
        t.config(state="disabled")

    def _contact_support(self):
        webbrowser.open(
            "mailto:samueldavidvida@gmail.com"
            "?subject=Docolco%20Product%20Manager%20-%20Soporte"
        )

    # ── Startup check ─────────────────────────────────────────────────────────

    def _startup_check(self):
        issues = []
        try:
            subprocess.run(["git", "--version"], capture_output=True, check=True)
        except (FileNotFoundError, subprocess.CalledProcessError):
            issues.append(
                "• Git no está instalado.\n"
                "  Descarga e instala Git para Windows (git-scm.com)."
            )
        if not PILLOW_OK:
            issues.append(
                "• Pillow no está instalado.\n"
                "  Corre tools/setup.bat primero."
            )
        if not OPENPYXL_OK:
            issues.append(
                "• openpyxl no está instalado.\n"
                "  Corre tools/setup.bat primero."
            )
        try:
            IMG_DIR.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            issues.append(f"• No se pudo crear img/productos/:\n  {e}")
        if not GENERATE_SCRIPT.exists():
            issues.append(
                "• No se encontró generate-products.py.\n"
                f"  Esperado en: {GENERATE_SCRIPT}"
            )
        if not EXCEL_PATH.exists():
            issues.append(
                "• No se encontró data/products.xlsx.\n"
                f"  Esperado en: {EXCEL_PATH}"
            )
        try:
            r = subprocess.run(["git", "remote", "-v"],
                               capture_output=True, text=True, cwd=str(REPO_ROOT))
            if not r.stdout.strip():
                issues.append("• El repositorio no tiene un remote de GitHub configurado.")
        except Exception:
            pass
        if issues:
            messagebox.showwarning(
                "Verificación inicial — problemas encontrados",
                "\n\n".join(issues)
                + "\n\nContacta a Samuel:\nsamueldavidvida@gmail.com\n+57 304 353 8450",
            )


if __name__ == "__main__":
    app = ImageManagerApp()
    app.mainloop()
