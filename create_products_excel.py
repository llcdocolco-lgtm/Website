"""
create_products_excel.py
Genera data/products.xlsx con el catálogo de Docolco LLC.
Ejecutar una sola vez para crear/recrear el Excel.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

os.makedirs('data', exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = 'Products'

# ── Colores ────────────────────────────────────────────────────────────────
RED        = PatternFill('solid', fgColor='C8202A')
BLUE       = PatternFill('solid', fgColor='1A3FA8')
YELLOW     = PatternFill('solid', fgColor='FFFDE7')
GRAY_CELL  = PatternFill('solid', fgColor='F0F0F0')
GRAY_ROW   = PatternFill('solid', fgColor='FAFAFA')
WHITE_ROW  = PatternFill('solid', fgColor='FFFFFF')
LEGEND_BG  = PatternFill('solid', fgColor='F5F5F5')

WHITE_BOLD = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
BLUE_BOLD  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
MUTED      = Font(name='Calibri', color='888888', size=9, italic=True)
NORMAL     = Font(name='Calibri', size=10)
SKU_FONT   = Font(name='Calibri', color='555555', size=10)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')

thin = Side(style='thin', color='DDDDDD')
BORDER = Border(bottom=thin)

# ── Fila 1 — Título ────────────────────────────────────────────────────────
ws.merge_cells('A1:G1')
ws['A1'] = 'DOCOLCO LLC — Product Catalog'
ws['A1'].fill      = RED
ws['A1'].font      = WHITE_BOLD
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 36

# ── Fila 2 — Instrucción ───────────────────────────────────────────────────
ws.merge_cells('A2:G2')
ws['A2'] = 'Edit yellow cells. Run  python generate-products.py  after saving.'
ws['A2'].fill      = LEGEND_BG
ws['A2'].font      = MUTED
ws['A2'].alignment = CENTER
ws.row_dimensions[2].height = 20

# ── Fila 3 — Leyenda ──────────────────────────────────────────────────────
ws.merge_cells('A3:C3')
ws['A3'] = 'Yellow = editable by client'
ws['A3'].fill      = LEGEND_BG
ws['A3'].font      = Font(name='Calibri', size=9, color='888888')
ws['A3'].alignment = LEFT
ws.merge_cells('D3:G3')
ws['D3'] = 'Gray = do not edit (SKU)'
ws['D3'].fill      = LEGEND_BG
ws['D3'].font      = Font(name='Calibri', size=9, color='888888')
ws['D3'].alignment = LEFT
ws.row_dimensions[3].height = 18

# ── Fila 4 — Headers ──────────────────────────────────────────────────────
headers = ['SKU', 'Name', 'Category', 'Unit Price', 'Box Price', 'Box Contents', 'Available']
widths  = [14,    32,      16,         14,            14,           20,              12]

for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.fill      = BLUE
    cell.font      = BLUE_BOLD
    cell.alignment = CENTER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[4].height = 26

# ── Datos — 13 productos ──────────────────────────────────────────────────
products = [
    (3870540, 'Soft Gloves S/M/L',                 'Protection', 1.65, 16.50, '12 Dozens',    'Y'),
    (3870541, 'Cling Wrap 300ft',                  'Packaging',  3.10, 31.00, '28 Rolls/Box', 'Y'),
    (761765,  'Liquid Detergent 1L',               'Cleaning',   2.30, 23.00, '12 Units/Box', 'Y'),
    (761710,  'Liquid Detergent 4L',               'Cleaning',   6.38, 63.80, '4 Units/Box',  'Y'),
    (761758,  'Cloth Softener 1L',                 'Cleaning',   2.30, 23.00, '12 Units/Box', 'Y'),
    (761772,  'Cloth Softener 4L',                 'Cleaning',   6.38, 63.80, '4 Units/Box',  'Y'),
    (761703,  'Dishwashing Liquid 0.5L',           'Cleaning',   1.48, 14.80, '24 Units/Box', 'Y'),
    (761789,  'Dishwashing Liquid 1L',             'Cleaning',   2.13, 21.30, '12 Units/Box', 'Y'),
    (761796,  'Floor Cleaner Lime Pomegranate 1L', 'Cleaning',   1.63, 16.30, '12 Units/Box', 'Y'),
    (761741,  'Floor Cleaner Lime Pomegranate 4L', 'Cleaning',   4.68, 46.80, '4 Units/Box',  'Y'),
    (761727,  'Floor Cleaner Sea Breeze 1L',       'Cleaning',   2.63, 26.30, '12 Units/Box', 'Y'),
    (761734,  'Floor Cleaner Sea Breeze 4L',       'Cleaning',   4.68, 46.80, '4 Units/Box',  'Y'),
    (7907173, 'Garbage Bags 3-pack',               'Waste',      3.30, 33.00, '10 Packs/Box', 'Y'),
]

for i, row_data in enumerate(products):
    row_num   = i + 5
    is_alt    = (i % 2 == 1)
    row_fill  = GRAY_ROW if is_alt else WHITE_ROW

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.alignment = CENTER if col_idx in (1, 3, 4, 5, 7) else LEFT
        cell.border    = BORDER

        if col_idx == 1:                 # SKU — gris, no editable
            cell.fill = GRAY_CELL
            cell.font = SKU_FONT
        elif col_idx in (4, 5):          # Precios — amarillo + formato
            cell.fill       = YELLOW
            cell.font       = NORMAL
            cell.number_format = '$#,##0.00'
        else:                            # Resto — amarillo, editable
            cell.fill = YELLOW
            cell.font = NORMAL

    ws.row_dimensions[row_num].height = 30

# ── Validaciones desplegables ─────────────────────────────────────────────
cat_dv = DataValidation(
    type='list',
    formula1='"Cleaning,Packaging,Protection,Waste"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle='Invalid category',
    error='Choose: Cleaning, Packaging, Protection or Waste'
)
avail_dv = DataValidation(
    type='list',
    formula1='"Y,N"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle='Invalid value',
    error='Enter Y or N'
)

cat_dv.sqref   = f'C5:C{4 + len(products)}'
avail_dv.sqref = f'G5:G{4 + len(products)}'
ws.add_data_validation(cat_dv)
ws.add_data_validation(avail_dv)

# ── Freeze pane debajo de headers ────────────────────────────────────────
ws.freeze_panes = 'A5'

# ── Guardar ──────────────────────────────────────────────────────────────
wb.save('data/products.xlsx')
print('OK  data/products.xlsx created successfully.')
print(f'   {len(products)} products — ready to edit and run generate-products.py')

